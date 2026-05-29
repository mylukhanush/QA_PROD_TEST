import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_kpi_excel(kpi_data, output_stream):
    """
    Generate a styled Excel spreadsheet from KPI results and write it to the output_stream.
    
    kpi_data is expected to be a list of dictionaries with keys:
    - category
    - parameter
    - loading_time
    - record_count
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance KPI Report"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    category_font = Font(name=font_family, size=10, bold=True, color="1E293B")
    parameter_font = Font(name=font_family, size=10, color="334155")
    value_font = Font(name=font_family, size=10, color="000000")
    
    # Fills
    header_fill = PatternFill(start_color="3B4252", end_color="3B4252", fill_type="solid") # Dark slate-blue
    category_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid") # Light steel-blue
    alternate_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light gray tint
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Alignments
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    category_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    parameter_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    value_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Build Header Rows
    # Row 1: Merged Category (A1:A2), Parameter (B1:B2), "Small" (C1:D1)
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:D1')
    
    ws['A1'] = "Category"
    ws['B1'] = "Parameter"
    ws['C1'] = "Small"
    
    # Row 2: Sub-headers
    ws['C2'] = "Loading Time (Sec)"
    ws['D2'] = "Number of Records"
    
    # Style Header Cells
    for r in range(1, 3):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.font = title_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            
    # Set Row Heights for headers
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    
    # Map raw parameters to categories and write data rows
    current_row = 3
    
    # Group kpi_data by category to handle merging
    categories_groups = []
    current_cat = None
    cat_items = []
    
    for item in kpi_data:
        cat = item.get("category", "")
        if cat != current_cat:
            if current_cat is not None:
                categories_groups.append((current_cat, cat_items))
            current_cat = cat
            cat_items = []
        cat_items.append(item)
    if current_cat is not None:
        categories_groups.append((current_cat, cat_items))
        
    for cat_name, items in categories_groups:
        start_row = current_row
        num_items = len(items)
        
        for idx, item in enumerate(items):
            # Fill parameter
            p_cell = ws.cell(row=current_row, column=2, value=item.get("parameter", ""))
            p_cell.font = parameter_font
            p_cell.alignment = parameter_align
            p_cell.border = thin_border
            
            # Fill loading time
            lt_val = item.get("loading_time", "")
            try:
                lt_val = float(lt_val) if lt_val not in ("", None) else ""
            except ValueError:
                pass
            lt_cell = ws.cell(row=current_row, column=3, value=lt_val)
            lt_cell.font = value_font
            lt_cell.alignment = value_align
            lt_cell.border = thin_border
            if isinstance(lt_val, float):
                lt_cell.number_format = '0.00'
                
            # Fill record count
            rc_val = item.get("record_count", "")
            if rc_val in ("", None, "None", "N/A"):
                rc_val = "-"
            try:
                if rc_val != "-":
                    rc_val = int(rc_val)
            except ValueError:
                pass
            
            # If the count is 0 or "0", display "No Data" as requested by user
            if rc_val == 0 or rc_val == "0":
                rc_val = "No Data"
                
            rc_cell = ws.cell(row=current_row, column=4, value=rc_val)
            rc_cell.font = value_font
            rc_cell.alignment = value_align
            rc_cell.border = thin_border
            
            # Zebra striping for parameters/values
            row_fill = alternate_fill if current_row % 2 == 0 else white_fill
            p_cell.fill = row_fill
            lt_cell.fill = row_fill
            rc_cell.fill = row_fill
            
            # Category column setup
            cat_cell = ws.cell(row=current_row, column=1)
            cat_cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 24
            current_row += 1
            
        # Merge Category cells
        end_row = current_row - 1
        if num_items > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            
        # Apply style to merged category cells
        for r in range(start_row, end_row + 1):
            c_cell = ws.cell(row=r, column=1)
            c_cell.fill = category_fill
            c_cell.font = category_font
            c_cell.alignment = category_align
            c_cell.border = thin_border
            
        # Set text on the top merged cell
        ws.cell(row=start_row, column=1, value=cat_name)
        
    # Auto-adjust column widths with minimum padding
    column_widths = {1: 25, 2: 42, 3: 20, 4: 20}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Save output
    wb.save(output_stream)

def generate_download_kpi_excel(kpi_data, output_stream):
    """
    Generate a styled Excel spreadsheet from Download KPI results and write it to the output_stream.
    
    kpi_data is expected to be a list of dictionaries with keys:
    - category
    - parameter
    - loading_time
    - record_count
    - expected_kpi
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Download Report"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    category_font = Font(name=font_family, size=10, bold=True, color="1E293B")
    parameter_font = Font(name=font_family, size=10, color="334155")
    value_font = Font(name=font_family, size=10, color="000000")
    
    # Fills
    header_fill = PatternFill(start_color="3B4252", end_color="3B4252", fill_type="solid") # Dark slate-blue
    category_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid") # Light steel-blue
    alternate_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light gray tint
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Alignments
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    category_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    parameter_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    value_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Single Header Row
    headers = ["Category", "Parameter", "Loading Time (Sec)", "Number of Records", "Expected KPI (Sec)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        
    ws.row_dimensions[1].height = 30
    
    current_row = 2
    
    # Group kpi_data by category to handle merging
    categories_groups = []
    current_cat = None
    cat_items = []
    
    for item in kpi_data:
        cat = item.get("category", "")
        if cat != current_cat:
            if current_cat is not None:
                categories_groups.append((current_cat, cat_items))
            current_cat = cat
            cat_items = []
        cat_items.append(item)
    if current_cat is not None:
        categories_groups.append((current_cat, cat_items))
        
    for cat_name, items in categories_groups:
        start_row = current_row
        num_items = len(items)
        
        for idx, item in enumerate(items):
            # Fill parameter
            p_cell = ws.cell(row=current_row, column=2, value=item.get("parameter", ""))
            p_cell.font = parameter_font
            p_cell.alignment = parameter_align
            p_cell.border = thin_border
            
            # Fill loading time
            lt_val = item.get("loading_time", "")
            try:
                lt_val = float(lt_val) if lt_val not in ("", None) else ""
            except ValueError:
                pass
            lt_cell = ws.cell(row=current_row, column=3, value=lt_val)
            lt_cell.font = value_font
            lt_cell.alignment = value_align
            lt_cell.border = thin_border
            if isinstance(lt_val, float):
                lt_cell.number_format = '0.00'
                
            # Fill record count
            rc_val = item.get("record_count", "")
            if rc_val in ("", None, "None", "N/A"):
                rc_val = "-"
            try:
                if rc_val != "-":
                    rc_val = int(rc_val)
            except ValueError:
                pass
            
            if rc_val == 0 or rc_val == "0":
                rc_val = "No Data"
                
            rc_cell = ws.cell(row=current_row, column=4, value=rc_val)
            rc_cell.font = value_font
            rc_cell.alignment = value_align
            rc_cell.border = thin_border
            
            # Fill expected KPI
            ek_val = item.get("expected_kpi", "")
            try:
                ek_val = float(ek_val) if ek_val not in ("", None) else ""
            except ValueError:
                pass
            ek_cell = ws.cell(row=current_row, column=5, value=ek_val)
            ek_cell.font = value_font
            ek_cell.alignment = value_align
            ek_cell.border = thin_border
            if isinstance(ek_val, float):
                ek_cell.number_format = '0.00'
            
            # Zebra striping for parameters/values
            row_fill = alternate_fill if current_row % 2 == 0 else white_fill
            p_cell.fill = row_fill
            lt_cell.fill = row_fill
            rc_cell.fill = row_fill
            ek_cell.fill = row_fill
            
            # Category column setup
            cat_cell = ws.cell(row=current_row, column=1)
            cat_cell.border = thin_border
            
            ws.row_dimensions[current_row].height = 24
            current_row += 1
            
        # Merge Category cells
        end_row = current_row - 1
        if num_items > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            
        # Apply style to merged category cells
        for r in range(start_row, end_row + 1):
            c_cell = ws.cell(row=r, column=1)
            c_cell.fill = category_fill
            c_cell.font = category_font
            c_cell.alignment = category_align
            c_cell.border = thin_border
            
        # Set text on the top merged cell
        ws.cell(row=start_row, column=1, value=cat_name)
        
    # Auto-adjust column widths with minimum padding
    column_widths = {1: 25, 2: 42, 3: 20, 4: 20, 5: 20}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    wb.save(output_stream)

